"""
Lead Scraper Backend — Flask + SocketIO + MongoDB
Scrapes: Gelbe Seiten + Google Maps (SerpAPI)
Filter:  ONLY businesses WITHOUT a website
"""

from flask import Flask, request, jsonify, send_file
from flask_socketio import SocketIO
from flask_cors import CORS
from pymongo import MongoClient, UpdateOne
from datetime import datetime
import requests
import re
import time
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

# ===== APP SETUP =====
app = Flask(__name__)
CORS(app, origins="*")
socketio = SocketIO(app, cors_allowed_origins="*", async_mode="threading")

# ===== CONFIG (edit these) =====
MONGO_URI = os.getenv("MONGO_URI", "mongodb://localhost:27017/")
SERPAPI_KEY = os.getenv("SERPAPI_KEY", "YOUR_SERPAPI_KEY")

# ===== MONGODB =====
client = MongoClient(MONGO_URI)
db = client["leadscout"]
leads_col = db["leads"]
scrape_logs_col = db["scrape_logs"]

# Indexes
leads_col.create_index([("Name", 1), ("City", 1)], unique=True, background=True)
leads_col.create_index("Category", background=True)
leads_col.create_index("City", background=True)
leads_col.create_index("LeadStatus", background=True)
leads_col.create_index("CreatedAt", background=True)

# ===== CATEGORIES (German trade) =====
CATEGORIES = [
    "klempner",
    "elektriker",
    "dachdecker",
    "sanitär",
    "gebäudereinigung",
    "fliesenleger",
    "fensterbau",
    "garagentor",
]

CATEGORY_LABELS = {
    "klempner": "Klempner",
    "elektriker": "Elektriker",
    "dachdecker": "Dachdecker",
    "sanitär": "Sanitär",
    "gebäudereinigung": "Gebäudereinigung",
    "fliesenleger": "Fliesenleger",
    "fensterbau": "Fensterbau",
    "garagentor": "Garten & Garagentor",
}

# ===== HEADERS =====
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "de-DE,de;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}


# ===== HELPERS =====
def extract_phones(text_list):
    """Extract up to 3 phone numbers from a list of strings, pad with 'None'."""
    phones = []
    pattern = re.compile(r"[\+0-9][\d\s\-\/\(\)]{5,}")
    for t in text_list:
        if t:
            found = pattern.findall(t)
            for f in found:
                cleaned = re.sub(r"\s+", " ", f).strip()
                if cleaned and cleaned not in phones:
                    phones.append(cleaned)
    while len(phones) < 3:
        phones.append("None")
    return phones[0], phones[1], phones[2]


def parse_opening_hours(raw):
    """Clean and return opening hours string."""
    if not raw:
        return "None"
    if isinstance(raw, dict):
        parts = []
        for day, hours in raw.items():
            parts.append(f"{day}: {hours}")
        return " | ".join(parts)
    if isinstance(raw, list):
        return " | ".join(str(h) for h in raw)
    cleaned = str(raw).strip()
    return cleaned if cleaned else "None"


# ===== SCRAPER: GELBE SEITEN =====
def scrape_gelbe_seiten(category, city, pages=4):
    """
    Scrape Gelbe Seiten for a category+city.
    Returns ONLY businesses with NO website.
    """
    leads = []
    base_url = "https://www.gelbeseiten.de/suche"

    for page_num in range(1, pages + 1):
        url = f"{base_url}/{category}/{city}"
        params = {"page": page_num} if page_num > 1 else {}

        try:
            res = requests.get(url, headers=HEADERS, params=params, timeout=15)
            res.raise_for_status()
        except Exception as e:
            print(f"[GelbeSeiten] Error fetching page {page_num}: {e}")
            continue

        soup = BeautifulSoup(res.text, "html.parser")
        results = soup.select("article[data-werbemittel-id], .mod-Treffer")

        if not results:
            # Try alternate selectors
            results = soup.select("[class*='Treffer']")

        for r in results:
            # ===== WEBSITE CHECK — SKIP if has website =====
            has_website = False
            links = r.select("a[href]")
            for link in links:
                href = link.get("href", "")
                if href.startswith("http") and "gelbeseiten" not in href and "tel:" not in href:
                    has_website = True
                    break
            website_tag = r.select_one("[class*='website'], [class*='Website'], a[class*='web']")
            if website_tag:
                has_website = True

            if has_website:
                continue

            # ===== EXTRACT FIELDS =====
            # Name
            name_el = (
                r.select_one(".mod-Treffer__name")
                or r.select_one("[class*='name']")
                or r.select_one("h2")
                or r.select_one("h3")
            )
            name = name_el.get_text(strip=True) if name_el else ""
            if not name:
                continue

            # Owner
            owner_el = r.select_one("[class*='inhaber'], [class*='owner'], [class*='kontakt']")
            owner = owner_el.get_text(strip=True) if owner_el else "None"

            # Phones
            phone_els = r.select("[class*='phone'], [class*='Phone'], [class*='tel'], a[href^='tel:']")
            raw_phones = []
            for el in phone_els:
                href = el.get("href", "")
                if href.startswith("tel:"):
                    raw_phones.append(href.replace("tel:", "").strip())
                else:
                    raw_phones.append(el.get_text(strip=True))

            phone1, phone2, phone3 = extract_phones(raw_phones)

            # Address
            addr_el = (
                r.select_one(".mod-Treffer__adresse")
                or r.select_one("[class*='adresse'], [class*='address']")
            )
            address = addr_el.get_text(strip=True) if addr_el else "None"

            # Opening hours
            hours_el = r.select_one("[class*='oeffnung'], [class*='hours'], [class*='zeit']")
            opening_hours = hours_el.get_text(strip=True) if hours_el else "None"

            # Extras — collect leftover text snippets
            extras_parts = []
            for tag in r.select("[class*='kategorie'], [class*='beschreibung'], [class*='tag'], [class*='badge']"):
                txt = tag.get_text(strip=True)
                if txt and len(txt) < 200:
                    extras_parts.append(txt)
            extras = " | ".join(extras_parts) if extras_parts else "None"

            leads.append({
                "Name": name,
                "Owner": owner,
                "Phone1": phone1,
                "Phone2": phone2,
                "Phone3": phone3,
                "Address": address,
                "Website": "None",
                "OpeningHours": opening_hours,
                "Extras": extras,
                "Source": "GelbeSeiten",
                "Category": CATEGORY_LABELS.get(category, category),
                "City": city.title(),
                "LeadStatus": "New",
                "CreatedAt": datetime.utcnow(),
            })

        time.sleep(1.2)  # polite delay

    return leads


# ===== SCRAPER: GOOGLE MAPS (SerpAPI) =====
def scrape_google_maps(category, city):
    """
    Scrape Google Maps via SerpAPI.
    Returns ONLY businesses with NO website.
    """
    leads = []

    if SERPAPI_KEY == "YOUR_SERPAPI_KEY":
        print("[GoogleMaps] SerpAPI key not configured — skipping.")
        return leads

    # Paginate through Google Maps results
    for start in [0, 20, 40]:
        try:
            res = requests.get(
                "https://serpapi.com/search.json",
                params={
                    "engine": "google_maps",
                    "q": f"{category} {city}",
                    "hl": "de",
                    "gl": "de",
                    "start": start,
                    "api_key": SERPAPI_KEY,
                },
                timeout=20,
            )
            data = res.json()
        except Exception as e:
            print(f"[GoogleMaps] Error: {e}")
            break

        places = data.get("local_results", [])
        if not places:
            break

        for place in places:
            # ===== WEBSITE CHECK — SKIP if has website =====
            if place.get("website"):
                continue

            # ===== EXTRACT FIELDS =====
            name = place.get("title", "")
            if not name:
                continue

            # Phones
            raw_phones = []
            if place.get("phone"):
                raw_phones.append(place["phone"])
            phone1, phone2, phone3 = extract_phones(raw_phones)

            # Address
            address = place.get("address", "None")

            # Owner (sometimes in "description" or "service_options")
            owner = "None"
            desc = place.get("description", "")
            if desc and len(desc) < 100:
                owner = desc

            # Opening hours
            hours_raw = place.get("hours") or place.get("operating_hours", {})
            opening_hours = parse_opening_hours(hours_raw)

            # Rating + reviews as extras
            extras_parts = []
            if place.get("rating"):
                extras_parts.append(f"Rating: {place['rating']}")
            if place.get("reviews"):
                extras_parts.append(f"Reviews: {place['reviews']}")
            if place.get("type"):
                extras_parts.append(f"Type: {place['type']}")
            if place.get("service_options"):
                for k, v in place["service_options"].items():
                    if v:
                        extras_parts.append(k)
            extras = " | ".join(extras_parts) if extras_parts else "None"

            leads.append({
                "Name": name,
                "Owner": owner,
                "Phone1": phone1,
                "Phone2": phone2,
                "Phone3": phone3,
                "Address": address,
                "Website": "None",
                "OpeningHours": opening_hours,
                "Extras": extras,
                "Source": "GoogleMaps",
                "Category": CATEGORY_LABELS.get(category, category),
                "City": city.title(),
                "LeadStatus": "New",
                "CreatedAt": datetime.utcnow(),
            })

        time.sleep(1.0)

    return leads


# ===== MERGE (deduplicate) =====
def merge_leads(lists):
    unique = {}
    for lead in [l for sub in lists for l in sub]:
        key = (lead["Name"].lower().strip(), lead["City"].lower().strip())
        if key not in unique:
            unique[key] = lead
    return list(unique.values())


# ===== SAVE (bulk upsert) =====
def save_leads(leads):
    if not leads:
        return 0
    ops = []
    for lead in leads:
        lead.setdefault("CreatedAt", datetime.utcnow())
        lead["UpdatedAt"] = datetime.utcnow()
        ops.append(
            UpdateOne(
                {"Name": lead["Name"], "City": lead["City"]},
                {
                    "$setOnInsert": {"CreatedAt": lead["CreatedAt"]},
                    "$set": {k: v for k, v in lead.items() if k != "CreatedAt"},
                },
                upsert=True,
            )
        )
    if ops:
        result = leads_col.bulk_write(ops, ordered=False)
        return result.upserted_count + result.modified_count
    return 0


# ===== BROADCAST =====
def broadcast(query=None):
    q = query or {}
    all_leads = list(leads_col.find(q, {"_id": 0}).sort("CreatedAt", -1).limit(500))
    socketio.emit("leads_update", all_leads)


# ===========================
# ===== API ROUTES ==========
# ===========================

@app.route("/api/categories")
def get_categories():
    return jsonify([{"key": k, "label": v} for k, v in CATEGORY_LABELS.items()])


@app.route("/api/leads")
def get_leads():
    # Build query from filters
    query = {}
    category = request.args.get("category")
    city = request.args.get("city")
    status = request.args.get("status")
    search = request.args.get("search")
    page = request.args.get("page", 1, type=int)
    page_size = request.args.get("pageSize", 50, type=int)

    if category:
        # match label or key
        label = CATEGORY_LABELS.get(category, category)
        query["Category"] = label
    if city:
        query["City"] = {"$regex": city, "$options": "i"}
    if status:
        query["LeadStatus"] = status
    if search:
        query["$or"] = [
            {"Name": {"$regex": search, "$options": "i"}},
            {"Address": {"$regex": search, "$options": "i"}},
            {"Phone1": {"$regex": search, "$options": "i"}},
        ]

    total = leads_col.count_documents(query)
    leads_list = list(
        leads_col.find(query, {"_id": 0})
        .sort("CreatedAt", -1)
        .skip((page - 1) * page_size)
        .limit(page_size)
    )

    return jsonify({"leads": leads_list, "total": total, "page": page, "pageSize": page_size})


@app.route("/api/scrape", methods=["POST"])
def scrape():
    body = request.get_json() or {}
    category = body.get("category", "elektriker")
    city = body.get("city", "berlin")

    socketio.emit("scrape_status", {"status": "running", "category": category, "city": city})

    try:
        gs_leads = scrape_gelbe_seiten(category, city, pages=4)
        gm_leads = scrape_google_maps(category, city)
        merged = merge_leads([gs_leads, gm_leads])
        saved = save_leads(merged)

        # Log
        scrape_logs_col.insert_one({
            "category": category,
            "city": city,
            "found": len(merged),
            "saved": saved,
            "gelbeseiten": len(gs_leads),
            "googlemaps": len(gm_leads),
            "timestamp": datetime.utcnow(),
        })

        broadcast()
        socketio.emit("scrape_status", {
            "status": "done",
            "found": len(merged),
            "saved": saved,
            "category": category,
            "city": city,
        })

        return jsonify({"ok": True, "found": len(merged), "saved": saved})

    except Exception as e:
        socketio.emit("scrape_status", {"status": "error", "message": str(e)})
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/update-status", methods=["POST"])
def update_status():
    data = request.get_json()
    leads_col.update_one(
        {"Name": data["Name"], "City": data["City"]},
        {"$set": {"LeadStatus": data["LeadStatus"], "UpdatedAt": datetime.utcnow()}},
    )
    broadcast()
    return jsonify({"ok": True})


@app.route("/api/delete-lead", methods=["POST"])
def delete_lead():
    data = request.get_json()
    leads_col.delete_one({"Name": data["Name"], "City": data["City"]})
    broadcast()
    return jsonify({"ok": True})


@app.route("/api/stats")
def get_stats():
    pipeline = [{"$group": {"_id": "$LeadStatus", "count": {"$sum": 1}}}]
    by_status = {item["_id"]: item["count"] for item in leads_col.aggregate(pipeline)}

    cat_pipeline = [{"$group": {"_id": "$Category", "count": {"$sum": 1}}}]
    by_category = {item["_id"]: item["count"] for item in leads_col.aggregate(cat_pipeline)}

    total = leads_col.count_documents({})

    return jsonify({
        "total": total,
        "byStatus": by_status,
        "byCategory": by_category,
    })


@app.route("/api/scrape-logs")
def get_scrape_logs():
    logs = list(scrape_logs_col.find({}, {"_id": 0}).sort("timestamp", -1).limit(20))
    return jsonify(logs)


@app.route("/api/export-excel")
def export_excel():
    query = {}
    category = request.args.get("category")
    city = request.args.get("city")
    status = request.args.get("status")
    if category:
        query["Category"] = CATEGORY_LABELS.get(category, category)
    if city:
        query["City"] = {"$regex": city, "$options": "i"}
    if status:
        query["LeadStatus"] = status

    leads = list(leads_col.find(query, {"_id": 0}))

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1a1a2e")
    header_align = Alignment(horizontal="center", vertical="center")

    headers = [
        "Name", "Owner", "Phone 1", "Phone 2", "Phone 3",
        "Address", "Website", "Opening Hours",
        "Category", "City", "Status", "Source", "Extras", "Created At"
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Rows
    for lead in leads:
        created = lead.get("CreatedAt")
        created_str = created.strftime("%Y-%m-%d %H:%M") if created else ""
        ws.append([
            lead.get("Name", ""),
            lead.get("Owner", "None"),
            lead.get("Phone1", "None"),
            lead.get("Phone2", "None"),
            lead.get("Phone3", "None"),
            lead.get("Address", "None"),
            lead.get("Website", "None"),
            lead.get("OpeningHours", "None"),
            lead.get("Category", ""),
            lead.get("City", ""),
            lead.get("LeadStatus", "New"),
            lead.get("Source", ""),
            lead.get("Extras", "None"),
            created_str,
        ])

    # Column widths
    col_widths = [30, 20, 18, 18, 18, 35, 10, 35, 20, 15, 12, 15, 40, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    file_path = "/tmp/leadscout_export.xlsx"
    wb.save(file_path)
    return send_file(file_path, as_attachment=True, download_name="leadscout_export.xlsx")


# ===== RUN =====
if __name__ == "__main__":
    print("🚀 LeadScout backend running on http://localhost:5000")
    socketio.run(app, host="0.0.0.0", port=5000, debug=False)
